"""
Модуль для сохранения результатов поиска в Excel
Формат: Дата | Маркетплейс | Бренд | Артикул | Название товара | Цена
"""
import os
import re
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# Excel файл в корне проекта
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'avito_prices.xlsx')

# Заголовки
HEADERS = [
    'Дата',
    'Маркетплейс',
    'Бренд',
    'Название товара',
    'Цена',
    'Ссылка на объявление'
]

# Известные бренды для автоопределения
KNOWN_BRANDS = [
    'Nike', 'Adidas', 'Puma', 'Reebok', 'New Balance', 'Asics',
    'Jordan', 'Converse', 'Vans', 'Fila', 'Under Armour', 'Skechers',
    'Salomon', 'Merrell', 'Timberland', 'Dr. Martens', 'Crocs',
    'Gucci', 'Louis Vuitton', 'Balenciaga', 'Versace', 'Prada',
    'Dior', 'Chanel', 'Hermes', 'Burberry', 'Zara', 'H&M',
    'Uniqlo', 'Levi\'s', 'Calvin Klein', 'Tommy Hilfiger', 'Lacoste',
    'Ralph Lauren', 'Hugo Boss', 'Armani', 'Diesel', 'Kappa',
    'The North Face', 'Columbia', 'Helly Hansen', 'Jack Wolfskin',
    'Apple', 'Samsung', 'Xiaomi', 'Huawei', 'Honor', 'Realme',
    'OnePlus', 'Google', 'Sony', 'LG', 'Motorola', 'Nokia',
    'iPhone', 'iPad', 'MacBook', 'AirPods', 'PlayStation', 'Xbox',
    'Nintendo', 'Dyson', 'Bosch', 'Makita', 'DeWalt',
]


def _detect_brand(title):
    """Определить бренд из названия товара"""
    if not title:
        return ''

    title_lower = title.lower()
    for brand in KNOWN_BRANDS:
        if brand.lower() in title_lower:
            return brand

    return ''


def _detect_article(title):
    """
    Попытка найти артикул в названии.
    Артикул обычно выглядит как: ABC-123, 12345, AB123CD
    """
    if not title:
        return ''

    # Паттерны артикулов
    patterns = [
        r'[Аа]рт\.?\s*:?\s*([A-Za-z0-9\-]+)',      # Арт: ABC123
        r'[Аа]ртикул\s*:?\s*([A-Za-z0-9\-]+)',       # Артикул: ABC123
        r'\b([A-Z]{2,}\-?\d{3,}[A-Za-z]*)\b',        # AB-12345, ABC123
        r'\b(\d{5,})\b',                               # 12345 (5+ цифр)
    ]

    for pattern in patterns:
        match = re.search(pattern, title)
        if match:
            return match.group(1)

    return ''


def _convert_relative_date(relative_str):
    """Конвертирует 'X часов назад' в реальную дату"""
    now = datetime.now()

    if not relative_str or relative_str.strip() == '':
        return now.strftime('%d-%m-%Y %H:%M')

    text = relative_str.lower().strip()

    # "5 минут назад"
    match = re.search(r'(\d+)\s*минут', text)
    if match:
        return (now - timedelta(minutes=int(match.group(1)))).strftime('%d-%m-%Y %H:%M')

    # "6 часов назад"
    match = re.search(r'(\d+)\s*час', text)
    if match:
        return (now - timedelta(hours=int(match.group(1)))).strftime('%d-%m-%Y %H:%M')

    # "3 дня назад"
    match = re.search(r'(\d+)\s*(день|дня|дней)', text)
    if match:
        return (now - timedelta(days=int(match.group(1)))).strftime('%d-%m-%Y %H:%M')

    # "вчера в 15:30"
    match = re.search(r'вчера.*?(\d{1,2}):(\d{2})', text)
    if match:
        yesterday = now - timedelta(days=1)
        return yesterday.replace(hour=int(match.group(1)), minute=int(match.group(2))).strftime('%d-%m-%Y %H:%M')

    # "сегодня в 15:30"
    match = re.search(r'сегодня.*?(\d{1,2}):(\d{2})', text)
    if match:
        return now.replace(hour=int(match.group(1)), minute=int(match.group(2))).strftime('%d-%m-%Y %H:%M')

    # "25 февраля 14:30"
    months = {
        'января': 1, 'февраля': 2, 'марта': 3, 'апреля': 4,
        'мая': 5, 'июня': 6, 'июля': 7, 'августа': 8,
        'сентября': 9, 'октября': 10, 'ноября': 11, 'декабря': 12
    }
    for month_name, month_num in months.items():
        match = re.search(rf'(\d{{1,2}})\s+{month_name}.*?(\d{{1,2}}):(\d{{2}})', text)
        if match:
            try:
                result = now.replace(month=month_num, day=int(match.group(1)),
                                     hour=int(match.group(2)), minute=int(match.group(3)))
                return result.strftime('%d-%m-%Y %H:%M')
            except ValueError:
                pass

    # "неделю назад"
    if 'недел' in text:
        match = re.search(r'(\d+)', text)
        weeks = int(match.group(1)) if match else 1
        return (now - timedelta(weeks=weeks)).strftime('%d-%m-%Y %H:%M')

    return f"{relative_str} ({now.strftime('%d-%m-%Y')})"


def _style_header(ws):
    """Стили заголовков"""
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_num, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.column_dimensions['A'].width = 20  # Дата
    ws.column_dimensions['B'].width = 15  # Маркетплейс
    ws.column_dimensions['C'].width = 18  # Бренд
    ws.column_dimensions['D'].width = 45  # Название товара
    ws.column_dimensions['E'].width = 18  # Цена
    ws.column_dimensions['F'].width = 50  # Ссылка на объявление


def _get_or_create_workbook():
    """Получить или создать Excel файл"""
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Мониторинг цен'
        _style_header(ws)
        wb.save(EXCEL_FILE)
        print(f"[EXCEL] Создан файл: {EXCEL_FILE}")

    return wb, ws


def save_results_to_excel(results, city_name='Казань', marketplace='Авито', brand=''):
    """
    Сохранить результаты в Excel.
    Формат: Дата | Маркетплейс | Бренд | Артикул | Название товара | Цена
    brand — то, что пользователь ввёл в поиск
    """
    wb, ws = _get_or_create_workbook()

    # Дата запроса = текущее время
    search_date = datetime.now().strftime('%d-%m-%Y %H:%M')

    data_alignment = Alignment(vertical='center', wrap_text=True)
    data_font = Font(name='Arial', size=10)
    brand_font = Font(name='Arial', size=10, bold=True)
    link_font = Font(name='Arial', size=10, color='0563C1', underline='single')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for product in results:
        title = product.get('title', '')
        url = product.get('url', '')

        row = [
            search_date,                                    # Дата
            marketplace,                                    # Маркетплейс
            brand,                                          # Бренд
            title,                                          # Название товара
            product.get('price', 'Цена не указана'),   # Цена
            url or '',                                      # Ссылка
        ]

        ws.append(row)

        last_row = ws.max_row
        for col_num in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=last_row, column=col_num)
            cell.alignment = data_alignment
            cell.border = thin_border
            cell.font = data_font

        # Бренд — жирным
        brand_cell = ws.cell(row=last_row, column=3)
        if brand_cell.value:
            brand_cell.font = brand_font

        # Ссылка на объявление — кликабельная
        if url and url != '#':
            link_cell = ws.cell(row=last_row, column=6)
            link_cell.hyperlink = url
            link_cell.value = 'Открыть объявление'
            link_cell.font = link_font

    wb.save(EXCEL_FILE)
    print(f"[EXCEL] Сохранено {len(results)} записей → {EXCEL_FILE}")

    return EXCEL_FILE


def get_excel_path():
    return EXCEL_FILE


def get_row_count():
    if not os.path.exists(EXCEL_FILE):
        return 0
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    count = ws.max_row - 1
    wb.close()
    return max(0, count)


if __name__ == '__main__':
    # Тест
    print("=== Тест нового формата Excel ===\n")

    test_data = [
        {
            'title': 'Nike Jordan Tatum 3 арт: DZ0115-001',
            'price': '15 000 ₽',
            'location': 'Казань',
            'date': '6 часов назад',
        },
        {
            'title': 'Кроссовки Adidas Ultraboost 22',
            'price': '8 500 ₽',
            'location': 'Казань',
            'date': 'вчера в 15:30',
        },
        {
            'title': 'Куртка зимняя мужская размер L',
            'price': '3 200 ₽',
            'location': 'Казань',
            'date': '3 дня назад',
        },
    ]

    path = save_results_to_excel(test_data, 'Казань', 'Авито')
    print(f"\nФайл: {path}")
    print(f"Записей: {get_row_count()}")
